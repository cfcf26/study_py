# 셀레니움
from selenium import webdriver
from selenium.webdriver.common.alert import Alert
# 엑셀
import openpyxl


# 셀레니움 - argument 설정 - 크롬드라이버 위치 설정
options = webdriver.ChromeOptions()
options.add_argument(
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36')  # user_argument를 사람처럼 수정함
driver = webdriver.Chrome('/Users/euncheal/chromedriver', options=options)


def login_admin():
    url = "https://admin3936.ableunion.com/"  # 어드민 주소 입력
    driver.get(url)
    driver.implicitly_wait(3)
    문어 = driver.find_element_by_xpath('//*[@id="frm"]/div[1]/input')
    문어.send_keys('냉장고')
    답어 = driver.find_element_by_xpath('//*[@id="frm"]/div[2]/input')
    답어.send_keys('너구리\n')
    # alert창 확인하기
    result = Alert(driver)
    result.accept()
    driver.implicitly_wait(3)


def find_import(day):
    # 배송대행내역관리 접속
    url = 'https://admin3936.ableunion.com/delivery/list.jsp'
    driver.get(url)
    # 정보 검색
    shop_name = driver.find_element_by_xpath('//*[@id="shop_name"]')
    shop_name.send_keys('-pack')
    등록일자 = driver.find_element_by_xpath('//*[@id="register_date"]')
    등록일자.send_keys(day)
    검색버튼 = driver.find_element_by_xpath(
        '//*[@id="page-wrapper"]/div[2]/div/div[2]/form/div/div[1]/div[7]/div[2]/input')
    검색버튼.click()
    driver.implicitly_wait(10)

    거래번호 = []
    사용자id = []
    shop_name = []
    product_name = []
    invoice = []
    상태 = []
    등록일자 = []
    위치 = []
    비고 = []

    # 1페이지 내용 리스트에 저장
    검색결과 = driver.find_element_by_tag_name(
        'tbody').find_elements_by_tag_name('tr')

    for i in 검색결과:
        거래번호.append(i.find_elements_by_tag_name('td')[0].text)
        사용자id.append(i.find_elements_by_tag_name('td')[1].text)
        shop_name.append(i.find_elements_by_tag_name('td')[2].text)
        product_name.append(i.find_elements_by_tag_name('td')[3].text)
        invoice.append(i.find_elements_by_tag_name('td')[4].text)
        상태.append(i.find_elements_by_tag_name('td')[5].text)
        등록일자.append(i.find_elements_by_tag_name('td')[6].text)
        위치.append(i.find_elements_by_tag_name('td')[7].text)
        비고.append(i.find_elements_by_tag_name('td')[8].text)

    # 2페이지 버튼 클릭
    driver.find_element_by_xpath(
        '//*[@id="page-wrapper"]/div[2]/div/div[2]/form/div/div[4]/div/div/button[2]').click()

    driver.implicitly_wait(10)
    # 2페이지 검색 결과 리스트에 저장
    검색결과2 = driver.find_element_by_tag_name(
        'tbody').find_elements_by_tag_name('tr')
    for j in 검색결과2:
        거래번호.append(j.find_elements_by_tag_name('td')[0].text)
        사용자id.append(j.find_elements_by_tag_name('td')[1].text)
        shop_name.append(j.find_elements_by_tag_name('td')[2].text)
        product_name.append(j.find_elements_by_tag_name('td')[3].text)
        invoice.append(j.find_elements_by_tag_name('td')[4].text)
        상태.append(j.find_elements_by_tag_name('td')[5].text)
        등록일자.append(j.find_elements_by_tag_name('td')[6].text)
        위치.append(j.find_elements_by_tag_name('td')[7].text)
        비고.append(j.find_elements_by_tag_name('td')[8].text)

    wb = openpyxl.Workbook()
    sheet = wb.active

    for i in range(1, len(거래번호)+1):
        sheet.cell(row=i, column=1).value = 거래번호[i-1]
        sheet.cell(row=i, column=2).value = 사용자id[i-1]
        sheet.cell(row=i, column=3).value = shop_name[i-1]
        sheet.cell(row=i, column=4).value = product_name[i-1]
        sheet.cell(row=i, column=5).value = invoice[i-1]
        sheet.cell(row=i, column=6).value = 상태[i-1]
        sheet.cell(row=i, column=7).value = 등록일자[i-1]
        sheet.cell(row=i, column=8).value = 위치[i-1]
        sheet.cell(row=i, column=9).value = 비고[i-1]
        sheet.cell(row=i, column=10).value = '=LEFT(E:E,5)'
    wb.save('test.xlsx')


login_admin()
find_import(20211210)
